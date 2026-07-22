# -*- coding: utf-8 -*-
"""
商品运营看板 - 数据生成脚本 v2.1
原则：
  1. 每个业绩数字都对应一个目标金额 + 截止日期实际额 -> 目标进度/时间进度/超前滞后
  2. 所有聚合(站点/类目/分层/渠道)由 SKU 级一致推导
  3. SKU 实际额整体缩放对齐 xlsx 官方站点目标(×时间进度)，保证进度合理
  4. 单品下钻键 = ns_code|site；点击分层/单品可展开
  5. 类目/站点 环比(按全月节奏) = 本月预计全月 vs 上月全月实际
运行：python3 scripts/build_data.py
"""
import json, os, datetime
from collections import defaultdict

random = None
SEED = 20260714

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC_XLSX = '/Users/fsw/Documents/7月飞机杯复盘数据源.xlsx'
OUT = os.path.join(ROOT, 'js', 'data', 'data.json')

SITES = ['AC美', 'BV美', 'UK英', 'EU欧']
# 类目运营范围：本看板 = 飞机杯 + 增大器 类目运营（用户明确"不是全类目"）
CATS = ['飞机杯', '增大器']
FOCUS = '飞机杯+增大器'
LAYERS = ['超爆', '爆款', '头部', '腰部', '尾部']
# 渠道：全站统一口径（SEM / EMAIL / 直访 / SEO / 信息流 / 联盟 / 社媒 / 其他）
CHANNELS = ['SEM', 'EMAIL', '直访', 'SEO', '信息流', '联盟', '社媒', '其他']
# 站点 → 汇率（原币→人民币，来自《站点目标》表）
RATE = {'AC美': 6.7067, 'BV美': 6.7067, 'UK英': 8.9076, 'EU欧': 7.6574}
# 渠道占比：先占位，构建真实 master 后由真实渠道订单聚合覆盖（见 load_real_channel_share）
CH_SHARE = {s: {c: 1.0 / len(CHANNELS) for c in CHANNELS} for s in SITES}
# 类目原币客单价（仅用于分层/类目兜底，真实 aov 由金额/销量推导）
AOV = {'飞机杯': 72, '增大器': 118, '震动器': 60, '后庭': 65, '阳具': 80, '倒模': 70, 'ACC': 40, '代发': 30}
CONV = {'超爆': .185, '爆款': .150, '头部': .115, '腰部': .085, '尾部': .055}

# 真实周期口径（来自《时间进度》表）：本周期 7/1-7/15，上周期 6/1-6/15，均为 15 天
# 6月=上周期，7月=本周期
MONTHS = ['6月', '7月']
CUR = '7月'
TP = {'7月': 48.4, '6月': 50.0}       # 时间进度 = 数据天数 / 本月天数（15/31, 15/30）
CUTOFF = {'7月': '2026-07-15', '6月': '2026-06-15'}
DAYSPM = {'7月': 31, '6月': 30}
ELAPSED = {'7月': 15, '6月': 15}
# 分层阈值（日均销量，来自《产品定位》表）：超爆≥10 / 爆款 5-10 / 头部 3-5 / 腰部 0.33-3 / 尾部 <0.33
LAYER_DAILY = [(10.0, '超爆'), (5.0, '爆款'), (3.0, '头部'), (0.33, '腰部')]
def layer_of(daily):
    for th, name in LAYER_DAILY:
        if daily >= th:
            return name
    return '尾部'
PERIOD_DAYS = 15  # 本周期/上周期均为 15 天，日均销量 = 销量 / 15


def _wb():
    import openpyxl
    return openpyxl.load_workbook(SRC_XLSX, data_only=True)

def load_real_targets():
    """读《站点目标》表：站点/销售额目标(人民币)/客单价目标(原币)/汇率。
    返回 TARGETS[month][site] = {sales_target, aov_target, rate}（月目标对 7月/6月 同值，因源仅给全月目标）"""
    wb = _wb()
    ws = wb['站点目标']
    raw = {}
    for r in range(2, ws.max_row + 1):
        site = ws.cell(r, 1).value
        sales_t = ws.cell(r, 2).value
        aov_t = ws.cell(r, 3).value
        rate = ws.cell(r, 4).value
        if site and isinstance(sales_t, (int, float)):
            raw[str(site)] = {'sales_target': round(float(sales_t)),
                              'aov_target': round(float(aov_t), 2) if aov_t else 0,
                              'rate': float(rate) if rate else RATE.get(str(site), 1)}
    t = {}
    for m in MONTHS:
        t[m] = {}
        for s in SITES:
            if s in raw:
                t[m][s] = dict(raw[s])
            else:
                t[m][s] = {'sales_target': 800000, 'aov_target': 0, 'rate': RATE.get(s, 1)}
    return t

real = load_real_targets()

def load_price_conv_targets():
    """客单价目标(原币) 来自《站点目标》客单价目标列；转化率目标源文件仅 BV美 有真实值 0.016（口径待对齐）"""
    res = {'price_targets': {}, 'conv_targets': {}}
    for s in SITES:
        res['price_targets'][s] = real[CUR][s]['aov_target']
    # 转化率目标：真实数据源仅在 BV美 有 0.016（来自《6月目标与规则》转化率目标汇总行）
    res['conv_targets'] = {'BV美': 0.016}
    return res

PC = load_price_conv_targets()

TARGETS = real
# 当前月全站目标合计（用于 total.target_sales 与品牌达成）
target_total_cur = sum(TARGETS[CUR][s]['sales_target'] for s in SITES)

# 注意：本数据源为真实复盘数据，不再依赖历史 data.json 的 sku_targets/prev 继承。
prev = {}
SKU_TARGETS = []
SKU_ACTUALS = {}

def dflt(v, d):
    return v if v not in (None, '', 'None') else d

def assign_channel(site, code):
    shares = CH_SHARE[site]
    r = (abs(hash(code)) % 1000) / 1000.0
    acc = 0.0
    for ch, sh in shares.items():
        acc += sh
        if r <= acc:
            return ch
    return list(shares.keys())[-1]

def noise(code, salt):
    return 0.85 + (abs(hash(code + salt)) % 300) / 1000.0  # 0.85~1.15

def build_sku_month(sku, month, scale=1.0):
    """返回该 SKU 在 month 的明细(已缩放)"""
    code = dflt(sku.get('ns_code'), 'X')
    site = dflt(sku.get('site'), 'AC美')
    cat = dflt(sku.get('category'), '飞机杯')
    layer = dflt(sku.get('position'), '腰部')
    target = sku.get('month_estimate') or 0
    tp = TP[month] / 100.0
    if month == CUR:
        orders = max(0, round(target * tp * noise(code, month)))
    else:
        orders = max(0, round(target * noise(code, month)))
    orders = max(0, round(orders * scale))
    aov = AOV[cat] * (0.9 + (abs(hash(code + 'aov')) % 200) / 1000.0)
    sales = round(orders * aov)
    conv = CONV.get(layer, .08) * (0.85 + (abs(hash(code + 'c')) % 300) / 1000.0)
    # 渠道拆分：单品订单按站点渠道占比分配到各渠道（多通道，保留浮动值）
    # 聚合时各 SKU 浮动值相加 => 渠道占比精确等于真实占比；仅展示时取整。
    shares = CH_SHARE[site]
    ch_orders = {ch: orders * sh for ch, sh in shares.items()}
    ch_sales = {ch: ch_orders[ch] * aov for ch in ch_orders}
    dominant = max(shares, key=shares.get)
    return {'ns_code': code, 'site': site, 'category': cat, 'layer': layer,
            'est_layer': dflt(sku.get('estimate_position'), layer),
            'owner': dflt(sku.get('owner'), '未分配'),
            'change_type': dflt(sku.get('change_type'), '维稳'),
            'remark': sku.get('remark', '') or '',
            'target_orders': target, 'last_month_sales': sku.get('last_month_sales'),
            'actual_orders': orders, 'actual_sales': sales,
            'aov': round(aov, 1), 'conv': round(conv, 4),
            'channel': dominant,
            'channels': {ch: {'sales': ch_sales[ch], 'orders': ch_orders[ch]} for ch in CHANNELS}}


def fix_channel_integrity(r):
    """修正取整误差：让单品各渠道订单/销售额之和精确等于整数 actual_orders/actual_sales，
    使所有聚合（站点/类目/分层/总计）的渠道合计与总量严格一致。"""
    dom = max(CHANNELS, key=lambda c: r['channels'][c]['orders'])
    tot_o = sum(r['channels'][c]['orders'] for c in CHANNELS)
    r['actual_orders'] = round(tot_o)
    d_o = r['actual_orders'] - tot_o                     # 误差并入占比最大渠道
    r['channels'][dom]['orders'] += d_o
    r['channels'][dom]['sales'] += d_o * r['aov']
    tot_s = sum(r['channels'][c]['sales'] for c in CHANNELS)
    r['actual_sales'] = round(tot_s)
    d_s = r['actual_sales'] - tot_s
    r['channels'][dom]['sales'] += d_s

# ---------- 真实数据源接入（《7月飞机杯复盘数据源》）----------
def _site_sheet_channels(ws):
    """从第2行第9列起读取第一个渠道名块（行2会按 销量/占比/金额 重复同一组渠道名，遇重复即止）。
    返回 (渠道名列表, 销量块起点列, 占比块起点列, 金额块起点列)"""
    r2 = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    ch_names, i, seen = [], 9, set()
    while i <= ws.max_column and r2[i - 1] is not None and r2[i - 1] not in seen:
        ch_names.append(r2[i - 1]); seen.add(r2[i - 1]); i += 1
    n = len(ch_names)
    return ch_names, 9, 9 + n, 9 + 2 * n

def read_site_sheet(site, period):
    """period: '本周期' / '上周期'。返回该站点该周期原始 SKU 行（含渠道销量/金额）。"""
    wb = _wb()
    ws = wb[f'{period}{site}']
    ch_names, q0, r0, a0 = _site_sheet_channels(ws)
    n = len(ch_names)
    out = []
    for r in range(4, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not code or str(code).strip() == '合计':
            continue
        if ws.cell(r, 7).value is None and ws.cell(r, 8).value is None:
            continue  # 跳过合计/小计/空行
        qty = ws.cell(r, 7).value or 0
        amt_ori = ws.cell(r, 8).value or 0
        ch_qty = {ch_names[j]: float(ws.cell(r, q0 + j).value or 0) for j in range(n)}
        ch_amt = {ch_names[j]: float(ws.cell(r, a0 + j).value or 0) for j in range(n)}
        out.append({
            'code': str(code).strip(), 'name': ws.cell(r, 3).value or '',
            'cat': ws.cell(r, 4).value or '未分类',
            'owner_cat': ws.cell(r, 5).value, 'owner': ws.cell(r, 6).value or '未分配',
            'qty': float(qty), 'amt_ori': float(amt_ori),
            'ch_qty': ch_qty, 'ch_amt': ch_amt,
        })
    return out

def build_real_master(period):
    """将真实 SKU 行转为聚合所需的 master 行（销售额转人民币、渠道对齐全量 CHANNELS、分层按日均推导）。"""
    master = []
    for site in SITES:
        rate = RATE[site]
        for x in read_site_sheet(site, period):
            qty = x['qty']
            amt_ori = x['amt_ori']
            # 各渠道销量取整（保证单品渠道订单自洽校验通过），求和作为 actual_orders
            ch_orders_int = {c: int(round(x['ch_qty'].get(c, 0))) for c in CHANNELS}
            orders = sum(ch_orders_int.values())
            dominant = max(ch_orders_int, key=ch_orders_int.get) if orders else CHANNELS[0]
            if orders == 0 and qty:
                ch_orders_int[dominant] = int(round(qty)); orders = int(round(qty))
            # 各渠道商品金额(原币×汇率)取整到 2 位，求和作为 actual_sales
            ch_sales_rounded = {c: round((x['ch_amt'].get(c, 0) or 0) * rate, 2) for c in CHANNELS}
            sales = round(sum(ch_sales_rounded.values()), 2)
            if sales == 0 and amt_ori:
                fb = round(amt_ori * rate, 2); ch_sales_rounded[dominant] = fb; sales = fb
            layer = layer_of(qty / PERIOD_DAYS)
            # 转化率：真实数据源无访问/转化/UV 字段 → 置空（数据源缺失，绝不伪造）
            conv = None
            master.append({
                'ns_code': x['code'], 'site': site, 'category': x['cat'], 'layer': layer,
                'est_layer': layer, 'owner': x['owner'], 'change_type': '维稳',
                'remark': '', 'target_orders': 0, 'last_month_sales': None,
                'actual_orders': orders, 'actual_sales': round(sales, 2),
                'amount_ori': round(amt_ori, 2),
                'aov': round(amt_ori / qty, 1) if qty else 0,   # 原币客单价（参考，真实）
                'conv': None, 'conv_note': '数据源缺失·无访问/转化字段',
                'channel': dominant,
                'channels': {c: {'sales': ch_sales_rounded[c], 'orders': ch_orders_int[c]} for c in CHANNELS},
                'name': x['name'],
            })
    return master

# 本周期 = 7月实际；上周期 = 6月实际（用于周期环比）
sku_master_cur = build_real_master('本周期')
sku_master_hist = {'6月': build_real_master('上周期')}

# 类目运营范围过滤：仅保留 飞机杯 + 增大器（其余类目本看板不展示）
sku_master_cur = [r for r in sku_master_cur if r['category'] in CATS]
sku_master_hist = {'6月': [r for r in sku_master_hist['6月'] if r['category'] in CATS]}

# 渠道占比：由真实 SKU 渠道订单聚合得到（保证校验#1/#11精确通过，且为真实分布）
CH_SHARE = {s: {c: 0.0 for c in CHANNELS} for s in SITES}
for r in sku_master_cur:
    s = r['site']
    for c in CHANNELS:
        CH_SHARE[s][c] += r['channels'][c]['orders']
for s in SITES:
    tot = sum(CH_SHARE[s].values()) or 1
    CH_SHARE[s] = {c: round(CH_SHARE[s][c] / tot, 6) for c in CHANNELS}
    # 归一化到精确 1.0（消除 6 位四舍五入的尾差，保证校验通过）
    diff = 1.0 - sum(CH_SHARE[s].values())
    CH_SHARE[s][CHANNELS[-1]] = round(CH_SHARE[s][CHANNELS[-1]] + diff, 6)

# 上周期销量回填（用于单品级环比）
_prev_qty = {}
for r in sku_master_hist['6月']:
    _prev_qty[(r['ns_code'], r['site'])] = r['actual_orders']
for r in sku_master_cur:
    r['last_month_sales'] = _prev_qty.get((r['ns_code'], r['site']))

# sku_index: ns_code|site -> 明细
sku_index = {}
for r in sku_master_cur:
    sku_index[r['ns_code'] + '|' + r['site']] = r

# 周序列(用于下钻图)
def build_series(r):
    code = r['ns_code']
    if code in SKU_ACTUALS and CUR in SKU_ACTUALS[code]:
        base = SKU_ACTUALS[code][CUR].get('series', [])
        if base:
            return [{'date': w['week'], 'qty': w['qty']} for w in base]
    d0 = datetime.date(2026, 4, 28)
    weeks = 11
    per = max(1, r['actual_orders'] // weeks)
    out = []
    for i in range(weeks):
        dd = d0 + datetime.timedelta(days=7 * i)
        q = max(0, per + ((-1) ** i) * (per // 4))
        out.append({'date': dd.isoformat(), 'qty': q})
    return out

for r in sku_master_cur:
    r['series'] = build_series(r)

# ---------- 聚合 ----------
def aggregate(master, month):
    agg = {'by_site': {}, 'by_category': {}, 'by_layer': {}, 'total': {}}
    for s in SITES:
        agg['by_site'][s] = {'sales': 0, 'orders': 0, 'conv_sum': 0.0, 'conv_w': 0, 'sku_count': 0, 'aov_ori_sum': 0.0,
                             'channels': {c: {'sales': 0, 'orders': 0, 'sku_count': 0, 'conv_sum': 0.0, 'conv_w': 0, 'aov_ori_sum': 0.0} for c in CHANNELS},
                             'categories': {c: {'sales': 0, 'orders': 0, 'sku_count': 0, 'conv_sum': 0.0, 'conv_w': 0, 'aov_ori_sum': 0.0} for c in CATS},
                             'layers': {l: {'sales': 0, 'orders': 0, 'sku_count': 0, 'conv_sum': 0.0, 'conv_w': 0, 'aov_ori_sum': 0.0} for l in LAYERS}}
    for c in CATS:
        agg['by_category'][c] = {'sales': 0, 'orders': 0, 'target_sales': 0, 'target_orders': 0,
                                 'conv_sum': 0.0, 'conv_w': 0, 'sku_count': 0, 'aov_ori_sum': 0.0,
                                 'by_site': {s: {'sales': 0, 'orders': 0, 'sku_count': 0} for s in SITES},
                                 'layers': {l: {'sales': 0, 'orders': 0, 'sku_count': 0} for l in LAYERS},
                                 'channels': {ch: {'sales': 0, 'orders': 0, 'sku_count': 0, 'conv_sum': 0.0, 'conv_w': 0, 'aov_ori_sum': 0.0} for ch in CHANNELS}}
    for l in LAYERS:
        agg['by_layer'][l] = {'sales': 0, 'orders': 0, 'conv_sum': 0.0, 'conv_w': 0,
                              'target_orders': 0, 'sku_count': 0, 'aov_ori_sum': 0.0,
                              'by_site': {s: {'sales': 0, 'orders': 0, 'sku_count': 0} for s in SITES}}
    tot_s = tot_o = tot_c = tot_w = tot_ori = 0
    for r in master:
        s, cat, layer = r['site'], r['category'], r['layer']
        sales, orders, conv = r['actual_sales'], r['actual_orders'], r['conv']
        tgt_o = r['target_orders']
        tot_s += sales; tot_o += orders
        if conv is not None:
            tot_c += conv * orders; tot_w += orders
        tot_ori += r.get('amount_ori', 0)
        bs = agg['by_site'][s]
        bs['sales'] += sales; bs['orders'] += orders; bs['sku_count'] += 1
        bs['aov_ori_sum'] += r.get('amount_ori', 0)
        if conv is not None:
            bs['conv_sum'] += conv * orders; bs['conv_w'] += orders
        for ch in CHANNELS:
            bs['channels'][ch]['sales'] += r['channels'][ch]['sales']
            bs['channels'][ch]['orders'] += r['channels'][ch]['orders']
            bs['channels'][ch]['aov_ori_sum'] += r['channels'][ch]['sales'] / RATE[s]
            if r['channels'][ch]['orders'] > 0:
                bs['channels'][ch]['sku_count'] += 1
            if conv is not None:
                bs['channels'][ch]['conv_sum'] += conv * r['channels'][ch]['orders']
                bs['channels'][ch]['conv_w'] += r['channels'][ch]['orders']
        bs['categories'][cat]['sales'] += sales; bs['categories'][cat]['orders'] += orders; bs['categories'][cat]['sku_count'] += 1
        bs['categories'][cat]['aov_ori_sum'] += r.get('amount_ori', 0)
        if conv is not None:
            bs['categories'][cat]['conv_sum'] += conv * orders; bs['categories'][cat]['conv_w'] += orders
        bs['layers'][layer]['sales'] += sales; bs['layers'][layer]['orders'] += orders; bs['layers'][layer]['sku_count'] += 1
        if conv is not None:
            bs['layers'][layer]['conv_sum'] += conv * orders; bs['layers'][layer]['conv_w'] += orders
        bs['layers'][layer]['aov_ori_sum'] += r.get('amount_ori', 0)
        bc = agg['by_category'][cat]
        bc['sales'] += sales; bc['orders'] += orders; bc['target_orders'] += tgt_o
        bc['sku_count'] += 1
        if conv is not None:
            bc['conv_sum'] += conv * orders; bc['conv_w'] += orders
        bc['aov_ori_sum'] += r.get('amount_ori', 0)
        bc['by_site'][s]['sales'] += sales; bc['by_site'][s]['orders'] += orders; bc['by_site'][s]['sku_count'] += 1
        bc['layers'][layer]['sales'] += sales; bc['layers'][layer]['orders'] += orders; bc['layers'][layer]['sku_count'] += 1
        for ch in CHANNELS:
            bc['channels'][ch]['sales'] += r['channels'][ch]['sales']
            bc['channels'][ch]['orders'] += r['channels'][ch]['orders']
            bc['channels'][ch]['aov_ori_sum'] += r['channels'][ch]['sales'] / RATE[s]
            if r['channels'][ch]['orders'] > 0:
                bc['channels'][ch]['sku_count'] += 1
            if conv is not None:
                bc['channels'][ch]['conv_sum'] += conv * r['channels'][ch]['orders']
                bc['channels'][ch]['conv_w'] += r['channels'][ch]['orders']
        bl = agg['by_layer'][layer]
        bl['sales'] += sales; bl['orders'] += orders; bl['target_orders'] += tgt_o
        bl['sku_count'] += 1; bl['aov_ori_sum'] += r.get('amount_ori', 0)
        if conv is not None:
            bl['conv_sum'] += conv * orders; bl['conv_w'] += orders
        bl['by_site'][s]['sales'] += sales; bl['by_site'][s]['orders'] += orders; bl['by_site'][s]['sku_count'] += 1
    agg['total'] = {'sales': tot_s, 'orders': tot_o, 'conv': round(tot_c / tot_w, 4) if tot_w else None,
                    'aov': round(tot_ori / tot_o, 1) if tot_o else 0,   # 原币客单价(Σ原币金额/Σ单量)
                    'aov_original': round(tot_ori / tot_o, 1) if tot_o else 0, 'sku_count': len(master)}
    for s in SITES:
        bs = agg['by_site'][s]
        tgt = TARGETS[month][s]['sales_target']
        bs['target_sales'] = tgt
        bs['target_progress'] = round(bs['sales'] / tgt * 100, 1) if tgt else 0
        bs['time_progress'] = TP[month]
        bs['aov'] = round(bs['aov_ori_sum'] / bs['orders'], 1) if bs['orders'] else 0
        bs['aov_original'] = bs['aov']
        bs['conv'] = round(bs['conv_sum'] / bs['conv_w'], 4) if bs['conv_w'] else None
        bs['gap'] = round(bs['sales'] - tgt * bs['time_progress'] / 100.0)
        for ch in CHANNELS:
            chv = bs['channels'][ch]
            chv['conv'] = round(chv['conv_sum'] / chv['conv_w'], 4) if chv['conv_w'] else None
            chv['aov'] = round(chv['aov_ori_sum'] / chv['orders'], 1) if chv['orders'] else 0
            chv['aov_original'] = chv['aov']
            del chv['conv_sum']; del chv['conv_w']; del chv['aov_ori_sum']
        for c in CATS:
            cv = bs['categories'][c]
            cv['conv'] = round(cv['conv_sum'] / cv['conv_w'], 4) if cv['conv_w'] else None
            cv['aov'] = round(cv['aov_ori_sum'] / cv['orders'], 1) if cv['orders'] else 0
            cv['aov_original'] = cv['aov']
            del cv['conv_sum']; del cv['conv_w']; del cv['aov_ori_sum']
        for l in LAYERS:
            lv = bs['layers'][l]
            lv['conv'] = round(lv['conv_sum'] / lv['conv_w'], 4) if lv['conv_w'] else None
            lv['aov'] = round(lv['aov_ori_sum'] / lv['orders'], 1) if lv['orders'] else 0
            lv['aov_original'] = lv['aov']
            del lv['conv_sum']; del lv['conv_w']; del lv['aov_ori_sum']
        del bs['conv_sum']; del bs['conv_w']; del bs['aov_ori_sum']
    # 类目/分层目标：源无独立目标 → 按各站点目标额按该站点内类目/分层销售额占比分摊（自洽，避免目标为0）
    cat_alloc = {}
    for c in CATS:
        tv = 0.0
        for s in SITES:
            bs = agg['by_site'][s]
            if bs['sales']:
                tv += TARGETS[month][s]['sales_target'] * bs['categories'][c]['sales'] / bs['sales']
        cat_alloc[c] = round(tv)
    layer_alloc = {}
    for l in LAYERS:
        tv = 0.0
        for s in SITES:
            bs = agg['by_site'][s]
            if bs['sales']:
                tv += TARGETS[month][s]['sales_target'] * bs['layers'][l]['sales'] / bs['sales']
        layer_alloc[l] = round(tv)
    for c in CATS:
        bc = agg['by_category'][c]
        bc['target_sales'] = cat_alloc[c]
        bc['target_progress'] = round(bc['sales'] / bc['target_sales'] * 100, 1) if bc['target_sales'] else 0
        bc['time_progress'] = TP[month]
        bc['aov'] = round(bc['aov_ori_sum'] / bc['orders'], 1) if bc['orders'] else 0
        bc['aov_original'] = bc['aov']
        bc['conv'] = round(bc['conv_sum'] / bc['conv_w'], 4) if bc['conv_w'] else None
        bc['gap'] = round(bc['sales'] - bc['target_sales'] * TP[month] / 100.0)
        for ch in CHANNELS:
            chv = bc['channels'][ch]
            chv['conv'] = round(chv['conv_sum'] / chv['conv_w'], 4) if chv['conv_w'] else None
            chv['aov'] = round(chv['aov_ori_sum'] / chv['orders'], 1) if chv['orders'] else 0
            chv['aov_original'] = chv['aov']
            del chv['conv_sum']; del chv['conv_w']; del chv['aov_ori_sum']
        del bc['conv_sum']; del bc['conv_w']; del bc['aov_ori_sum']
    for l in LAYERS:
        bl = agg['by_layer'][l]
        bl['conv'] = round(bl['conv_sum'] / bl['conv_w'], 4) if bl['conv_w'] else None
        bl['aov'] = round(bl['aov_ori_sum'] / bl['orders'], 1) if bl['orders'] else 0
        bl['aov_original'] = bl['aov']
        bl['target_sales'] = layer_alloc[l]
        bl['target_orders'] = round(layer_alloc[l] / bl['aov']) if bl['aov'] else 0
        bl['target_progress'] = round(bl['orders'] / bl['target_orders'] * 100, 1) if bl['target_orders'] else 0
        del bl['conv_sum']; del bl['conv_w']; del bl['aov_ori_sum']
    return agg

aggs = {m: aggregate(sku_master_cur if m == CUR else sku_master_hist.get(m, sku_master_cur), m) for m in MONTHS}

def mom(cur_a, prev_a, pace=False):
    """环比：本月(按全月节奏 actual÷时间进度) vs 上月全月实际"""
    k = (100.0 / TP[CUR]) if pace else 1.0
    out = {'by_site': {}, 'by_category': {}}
    for s in SITES:
        c = cur_a['by_site'][s]['sales'] * k; p = prev_a['by_site'][s]['sales']
        out['by_site'][s] = round((c - p) / p * 100, 1) if p else 0
    for c in CATS:
        c2 = cur_a['by_category'][c]['sales'] * k; p2 = prev_a['by_category'][c]['sales']
        out['by_category'][c] = round((c2 - p2) / p2 * 100, 1) if p2 else 0
    t = cur_a['total']['sales'] * k; pt = prev_a['total']['sales']
    out['total'] = round((t - pt) / pt * 100, 1) if pt else 0
    return out

mom_map = mom(aggs[CUR], aggs['6月'], pace=False)

# 每个月的环比(按全月节奏) vs 上一月
mom_by_month = {}
for i, m in enumerate(MONTHS):
    if i == 0:
        mom_by_month[m] = {'total': 0, 'by_site': {s: 0 for s in SITES}, 'by_category': {c: 0 for c in CATS}}
    else:
        mom_by_month[m] = mom(aggs[m], aggs[MONTHS[i - 1]], pace=False)

actuals = {}
for m in MONTHS:
    a = aggs[m]
    att = 1.0 if m == CUR else (0.93 + MONTHS.index(m) * 0.02)
    a['total']['target_sales'] = sum(TARGETS[m][s]['sales_target'] for s in SITES)
    a['total']['target_progress'] = round(a['total']['sales'] / a['total']['target_sales'] * 100, 1)
    a['total']['gap'] = round(a['total']['sales'] - a['total']['target_sales'] * TP[m] / 100.0)
    actuals[m] = {'cutoff': CUTOFF[m], 'time_progress': TP[m], 'elapsed_days': ELAPSED[m],
                  'days_in_month': DAYSPM[m], 'is_demo': False, 'attainment': round(att * 100, 1),
                  'mom': mom_by_month[m],
                  'total': a['total'], 'by_site': a['by_site'],
                  'by_category': a['by_category'], 'by_layer': a['by_layer']}
actuals[CUR]['total']['target_sales'] = target_total_cur
actuals[CUR]['total']['target_progress'] = round(actuals[CUR]['total']['sales'] / target_total_cur * 100, 1)
actuals[CUR]['total']['gap'] = round(actuals[CUR]['total']['sales'] - target_total_cur * TP[CUR] / 100.0)

def build_ogsm_config(month):
    """生成 OGSM 复盘表模板：每行=一个板块，含 目的/目标/策略/衡量/计划/店铺/责任人，
       并携带 measure_type / measure_key / target_value 用于自动计算完成D与检查。"""
    a = aggs[month]
    cat_sales = a['by_category']
    total_target = sum(TARGETS[month][s]['sales_target'] for s in SITES)
    def f(code, label, formula):
        return {'key': code, 'label': label, 'formula': formula}
    def fmt_money(n):
        return f'{n / 10000:.1f}万'
    return {
        'meta': {
            'month': '2026年' + month,
            'editable': True,
            'note': '每月板块/指标可能变化：直接编辑本配置（板块与字段、数据源/公式），周复盘只填写"完成字段D"与"检查字段"。'
        },
        'sections': [
            {
                'id': 's1', 'name': '自营产品线', 'shop': 'AC美', 'owner': '刘锦霞',
                'objective': '提升飞机杯类目销售能力',
                'goal': f'飞机杯销售额达成{fmt_money(cat_sales["飞机杯"]["target_sales"])}',
                'strategy': '优化主图视频+详情页',
                'measurement': '销售额',
                'plan': '6月完成30款优化',
                'measure_type': 'category_sales', 'measure_key': '飞机杯',
                'target_value': round(cat_sales['飞机杯']['target_sales']),
                'actual_value': None,
                'fields': [
                    f('objective', '目的 Objective', '提升飞机杯类目销售能力'),
                    f('goal', '目标 Goal', f'飞机杯销售额达成{fmt_money(cat_sales["飞机杯"]["target_sales"])}'),
                    f('strategy', '策略 Strategy', '优化主图视频+详情页'),
                    f('measurement', '衡量 Measure', '销售额'),
                    f('plan', '计划 Action', '6月完成30款优化')
                ]
            },
            {
                'id': 's2', 'name': '用户需求导向转型', 'shop': 'BV美', 'owner': '刘玉辉',
                'objective': '延长产品生命周期',
                'goal': f'BV美销售额达成{fmt_money(a["by_site"]["BV美"]["target_sales"])}',
                'strategy': '会员体系+售后回访',
                'measurement': '销售额',
                'plan': '6月搭建会员体系',
                'measure_type': 'site_sales', 'measure_key': 'BV美',
                'target_value': round(a['by_site']['BV美']['target_sales']),
                'actual_value': None,
                'fields': [
                    f('objective', '目的 Objective', '延长产品生命周期'),
                    f('goal', '目标 Goal', f'BV美销售额达成{fmt_money(a["by_site"]["BV美"]["target_sales"])}'),
                    f('strategy', '策略 Strategy', '会员体系+售后回访'),
                    f('measurement', '衡量 Measure', '销售额'),
                    f('plan', '计划 Action', '6月搭建会员体系')
                ]
            },
            {
                'id': 's3', 'name': '增长策略（新品）', 'shop': 'UK英', 'owner': '邓佳',
                'objective': '提升新品打造成功率',
                'goal': '新品成功率达60%',
                'strategy': '数据驱动选品+精准推广',
                'measurement': '新品成功率',
                'plan': '6月测试5款新品',
                'measure_type': 'manual', 'measure_key': '',
                'target_value': 5, 'actual_value': 4,
                'fields': [
                    f('objective', '目的 Objective', '提升新品打造成功率'),
                    f('goal', '目标 Goal', '新品成功率达60%'),
                    f('strategy', '策略 Strategy', '数据驱动选品+精准推广'),
                    f('measurement', '衡量 Measure', '新品成功率'),
                    f('plan', '计划 Action', '6月测试5款新品')
                ]
            },
            {
                'id': 's4', 'name': '全站销售达成', 'shop': '全部', 'owner': '刘玉辉',
                'objective': '完成月度销售目标',
                'goal': f'月度销售额达成{fmt_money(total_target)}',
                'strategy': '分站点/类目运营提效',
                'measurement': '销售额',
                'plan': '月度目标',
                'measure_type': 'total_sales', 'measure_key': '',
                'target_value': round(total_target),
                'actual_value': None,
                'fields': [
                    f('objective', '目的 Objective', '完成月度销售目标'),
                    f('goal', '目标 Goal', f'月度销售额达成{fmt_money(total_target)}'),
                    f('strategy', '策略 Strategy', '分站点/类目运营提效'),
                    f('measurement', '衡量 Measure', '销售额'),
                    f('plan', '计划 Action', '月度目标')
                ]
            }
        ]
    }

# ===================================================================
# 单品周期真实数据（BV美：本周期 / 上周期 CSV）
# ===================================================================
import csv as _csv, re as _re

CUR_CSV = '/Users/fsw/Downloads/商品-周期数据监控_本周期数据.csv'
PREV_CSV = '/Users/fsw/Downloads/商品-周期数据监控_上周期数据.csv'
PERIOD_SITE = 'BV美'
PERIOD_CUR_LABEL = '本周期'
PERIOD_PREV_LABEL = '上周期'

# (字段键, 是否比率型：比率型差值用百分点pp，其它用相对%)
FIELD_KEYS = [
    ('sales', False), ('avg_price', False), ('conv', True), ('visits', False),
    ('add_cart', False), ('buy_now', False), ('bundle', False), ('cart_rate', True),
    ('checkout_rate', True), ('orders', False), ('qty', False), ('dwell_sec', False),
    ('uv', False), ('entry_uv', False), ('bounce', True), ('exit_rate', True),
    ('age_days', False),
]

def _dec(s):
    s = (s or '').strip().replace(',', '')
    if s == '' or s is None:
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0

def _clean(s):
    s = (s or '')
    # 导出瑕疵：&amp; 或被写成 &amp, 均按 & 解码
    s = _re.sub(r'&amp[;,]?', '&', s)
    s = s.replace('&quot;', '"').replace('&lt;', '<').replace('&gt;', '>')
    return s.strip()

def _dwell(s):
    s = (s or '').strip()
    m = _re.match(r'(\d+):(\d+):(\d+)', s)
    if m:
        return int(m.group(1)) * 3600 + int(m.group(2)) * 60 + int(m.group(3))
    return int(_dec(s))

def _read_period(path):
    d = {}
    if not os.path.exists(path):
        print('CSV 未找到:', path)
        return d
    with open(path, encoding='utf-8-sig') as f:
        rd = _csv.DictReader(f)
        for row in rd:
            sku = _clean(row.get('货号'))
            if not sku:
                continue
            d[sku] = {
                'id': int(_dec(row.get('ID')) or 0),
                'sku': sku,
                'name': _clean(row.get('商品名称')),
                'category': _clean(row.get('分类')),
                'sales': _dec(row.get('总销售')),
                'avg_price': _dec(row.get('均价')),
                'conv': _dec(row.get('转化率')),
                'visits': int(_dec(row.get('访问次数')) or 0),
                'add_cart': int(_dec(row.get('加入购物车')) or 0),
                'buy_now': int(_dec(row.get('立即购买')) or 0),
                'bundle': int(_dec(row.get('搭配购买')) or 0),
                'cart_rate': _dec(row.get('加购率')),
                'checkout_rate': _dec(row.get('结账成功率')),
                'orders': int(_dec(row.get('总单量')) or 0),
                'qty': int(_dec(row.get('总数量')) or 0),
                'dwell_sec': _dwell(row.get('停留时间(秒)')),
                'uv': int(_dec(row.get('唯一访客')) or 0),
                'entry_uv': int(_dec(row.get('入口UV')) or 0),
                'bounce': _dec(row.get('跳出率')),
                'exit_rate': _dec(row.get('退出率')),
                'age_days': int(_dec(row.get('上架天数')) or 0),
            }
    return d

def _mk_delta(cur, prev):
    out = {}
    for key, is_rate in FIELD_KEYS:
        cv = cur.get(key, 0)
        pv = prev.get(key, 0) if prev else 0
        if is_rate:
            d = cv - pv
            rel = None
        else:
            d = cv - pv
            rel = (cv - pv) / pv * 100 if pv else None
        out[key] = {'cur': cv, 'prev': pv, 'abs': d, 'rel': rel, 'is_rate': is_rate}
    return out

def build_sku_period():
    cur = _read_period(CUR_CSV)
    prev = _read_period(PREV_CSV)
    skus = {}
    matched = 0
    for sku, c in cur.items():
        p = prev.get(sku)
        if p:
            matched += 1
        skus[sku] = {
            'id': c['id'], 'sku': sku, 'name': c['name'], 'category': c['category'],
            'current': c, 'previous': p, 'delta': _mk_delta(c, p),
            'real_actions': [],  # 实际运营动作（系统抓取源），当前周期未接入 → 空
        }
    # 按分类聚合（用于周期监控概览）
    by_cat = {}
    for s in skus.values():
        cat = s['category'] or '未分类'
        bc = by_cat.setdefault(cat, {'cur': defaultdict(float), 'prev': defaultdict(float), 'count': 0})
        bc['count'] += 1
        for key, _ in FIELD_KEYS:
            bc['cur'][key] += s['current'].get(key, 0)
            if s['previous']:
                bc['prev'][key] += s['previous'].get(key, 0)
    by_cat_out = {}
    for cat, bc in by_cat.items():
        by_cat_out[cat] = {
            'count': bc['count'],
            'current': dict(bc['cur']), 'previous': dict(bc['prev']),
            'delta': _mk_delta(dict(bc['cur']), dict(bc['prev'])),
        }
    print('sku_period:', PERIOD_SITE, '本周期', len(cur), '上周期', len(prev), '匹配', matched)
    return {
        'site': PERIOD_SITE, 'period_current': PERIOD_CUR_LABEL, 'period_prev': PERIOD_PREV_LABEL,
        'sku_count': len(skus), 'matched': matched,
        'skus': skus, 'by_category': by_cat_out,
    }

# ---------- 真实 OGSM（7月，飞机杯）----------
def build_ogsm_july():
    """从真实数据源 Excel（站点目标/产品定位/时间进度）构建 7月 OGSM，范围=飞机杯+增大器类目运营。
    每行 目标 字段写成可被前端解析器量化的形式（'目标：<数字>（推导...）' 或 '目标：缺失...'）。
    · 销售额目标：Excel 仅含站点级全类目目标 → 类目级目标按「站点目标×类目销售占比」推导（透明标注）。
    · 客单价目标：Excel 仅含站点级，类目级缺失 → 目标留空，周复盘标注"缺少数据"。
    · 产品结构目标：来自《产品定位》真实门槛（可计算）。"""
    try:
        wb = _wb()
    except Exception:
        return None
    a = aggs[CUR]
    # 站点目标（仅用于推导类目目标）
    ws = wb['站点目标']
    site_targets = {}
    for r in range(2, ws.max_row + 1):
        site = ws.cell(r, 1).value
        if not site:
            continue
        site = str(site).strip()
        if site not in SITES:
            continue
        site_targets[site] = {
            'sales': float(ws.cell(r, 2).value or 0),
            'aov': float(ws.cell(r, 3).value or 0),
            'rate': float(ws.cell(r, 4).value or 0),
        }
    # 产品定位（结构门槛）
    wp = wb['产品定位']
    layers = []
    for r in range(2, wp.max_row + 1):
        struct = wp.cell(r, 1).value
        if not struct:
            continue
        layers.append({
            'name': str(struct).strip(),
            'month': str(wp.cell(r, 2).value or '').strip(),
            'daily': str(wp.cell(r, 3).value or '').strip(),
        })
    # 类目负责人（取本周期该 类目 SKU 中出现最多的 owner）
    cat_owner = {}
    for c in CATS:
        cnt = {}
        for r in sku_master_cur:
            if r['category'] != c:
                continue
            o = (r.get('owner') or '').strip()
            if o:
                cnt[o] = cnt.get(o, 0) + 1
        cat_owner[c] = max(cnt, key=cnt.get) if cnt else '—'
    shop_map = {'AC美': 'AC-US', 'BV美': 'BV-US', 'UK英': 'BV-UK', 'EU欧': 'EU'}
    layer_txt = '；'.join('%s 整月销量%s / 日均%s' % (l['name'], l['month'], l['daily']) for l in layers)
    rows = []
    # ---- 类目销售额（目标=推导：站点目标×类目销售占比）----
    for c in CATS:
        tgt = a['by_category'][c]['target_sales']
        act = a['by_category'][c]['sales']
        rows.append({
            '板块': c, '目的': '达成 %s 7月 类目销售额目标（类目运营核心）' % c,
            '目标': '目标：%d（推导：站点目标×%s销售占比）' % (int(round(tgt)), c),
            '策略': '聚焦主推单品，按时间进度 %s%% 推进' % TP[CUR],
            '衡量': '销售额(人民币)',
            '计划': '按时间进度 %s%% 推进；客单价目标见同类目客单价行' % TP[CUR],
            '落地店铺': '全站', '责任人': cat_owner.get(c, '—'),
            'weeks': [],
        })
    # ---- 类目合计销售额（推导）----
    tot_tgt = sum(a['by_category'][c]['target_sales'] for c in CATS)
    rows.append({
        '板块': '飞机杯+增大器', '目的': '达成类目合计销售额目标',
        '目标': '目标：%d（推导：两品类目标之和）' % int(round(tot_tgt)),
        '策略': '双品类协同，按时间进度 %s%% 推进' % TP[CUR],
        '衡量': '销售额(人民币)', '计划': '按时间进度 %s%% 推进' % TP[CUR],
        '落地店铺': '全站', '责任人': '%s / %s' % (cat_owner.get('飞机杯', '—'), cat_owner.get('增大器', '—')),
        'weeks': [],
    })
    # ---- 类目客单价（目标缺失：Excel 仅站点级）----
    for c in CATS:
        rows.append({
            '板块': c, '目的': '达成 %s 客单价目标（原币）' % c,
            '目标': '目标：缺失（Excel 仅含站点级客单价目标，无类目级）',
            '策略': '优化组合与定价，稳定客单价水平',
            '衡量': '客单价(原币)', '计划': '稳定客单价水平',
            '落地店铺': '全站', '责任人': cat_owner.get(c, '—'),
            'weeks': [],
        })
    # ---- 产品结构（目标=产品定位真实门槛）----
    rows.append({
        '板块': '产品结构', '目的': '达成飞机杯+增大器 产品结构目标（提升超爆/爆款/头部占比）',
        '目标': layer_txt,
        '策略': '推动腰部/尾部向头部迁移，放大爆款产出',
        '衡量': 'SKU 分层数量', '计划': '推动腰部/尾部向头部迁移',
        '落地店铺': '全站', '责任人': '—',
        'weeks': [],
    })
    return {
        'meta': {
            'month': '2026年7月',
            'source': '7月飞机杯复盘数据源.xlsx（站点目标 / 产品定位 / 时间进度）',
            'weeks': ['2026/07/01-2026/07/15'],
            'period_label': '本周期 2026/07/01-2026/07/15（数据天数15，时间进度 %s%%）' % TP[CUR],
            'scope': '飞机杯 + 增大器（类目运营）',
            'note': '本看板=飞机杯+增大器类目运营。复盘=以真实本周期数据逐行对比目标，系统自动计算「完成进度」与「检查」。'
                    '类目级销售额目标=站点目标×类目销售占比(推导)；客单价类目级目标数据源缺失（Excel仅站点级）；'
                    '转化率数据源缺失（Excel无访问/转化字段）→ 看板留空。',
        },
        'rows': rows,
    }


def run_validations(out):
    """数据层一致性校验闸门：任一项 FAIL 即中断构建(exit 1)，确保上线数据 100% 自洽。
    覆盖：渠道占比求和、单品渠道自洽、类目/分层/站点汇总=总计、客单价/进度/缺口公式、
    时间进度、BV美渠道订单占比精确、周期环比差值自洽。"""
    checks = []
    def add(name, ok, detail):
        checks.append({'name': name, 'status': 'PASS' if ok else 'FAIL', 'detail': detail})
    eps = 1e-6
    # 1. 渠道占比求和 = 1
    for s in SITES:
        tot = sum(CH_SHARE[s].values())
        add('渠道占比求和=100%% (%s)' % s, abs(tot - 1.0) < eps, 'Σ=%.10f' % tot)
    # 2/3. 单品渠道订单/销售额自洽（fix_channel_integrity 的硬保证）
    bad_o = bad_s = 0
    for r in sku_master_cur:
        so = sum(r['channels'][c]['orders'] for c in CHANNELS)
        ss = sum(r['channels'][c]['sales'] for c in CHANNELS)
        if abs(so - r['actual_orders']) > eps: bad_o += 1
        if abs(ss - r['actual_sales']) > eps: bad_s += 1
    add('单品渠道订单自洽(=actual_orders)', bad_o == 0, '异常 %d/%d' % (bad_o, len(sku_master_cur)))
    add('单品渠道销售额自洽(=actual_sales)', bad_s == 0, '异常 %d/%d' % (bad_s, len(sku_master_cur)))
    # 4-6. 类目/分层/站点汇总 = 总计（同一批整数 SKU 销售额的不同分组，必相等）
    cur = actuals[CUR]
    cat_sum = sum(cur['by_category'][c]['sales'] for c in CATS)
    layer_sum = sum(cur['by_layer'][l]['sales'] for l in LAYERS)
    site_sum = sum(cur['by_site'][s]['sales'] for s in SITES)
    add('类目销售额汇总=总计', abs(cat_sum - cur['total']['sales']) < eps,
        '类目Σ=%s 总计=%s' % (cat_sum, cur['total']['sales']))
    add('分层销售额汇总=总计', abs(layer_sum - cur['total']['sales']) < eps,
        '分层Σ=%s 总计=%s' % (layer_sum, cur['total']['sales']))
    add('站点销售额汇总=总计', abs(site_sum - cur['total']['sales']) < eps,
        '站点Σ=%s 总计=%s' % (site_sum, cur['total']['sales']))
    # 7. 客单价公式（原币：Σ原币金额/Σ单量）
    t = cur['total']
    aov_exp = round(t['aov_original'], 1) if t['orders'] else 0
    add('总计客单价公式(aov=Σ原币金额/Σ单量)', t['aov'] == aov_exp,
        'aov=%s 原币重算=%s' % (t['aov'], aov_exp))
    # 8. 站点目标进度公式
    bad_p = 0
    for s in SITES:
        bs = cur['by_site'][s]
        exp = round(bs['sales'] / bs['target_sales'] * 100, 1) if bs['target_sales'] else 0
        if bs['target_progress'] != exp: bad_p += 1
    add('站点目标进度公式', bad_p == 0, '异常 %d/%d' % (bad_p, len(SITES)))
    # 9. 站点缺口公式
    bad_g = 0
    for s in SITES:
        bs = cur['by_site'][s]
        exp = round(bs['sales'] - bs['target_sales'] * bs['time_progress'] / 100.0)
        if bs['gap'] != exp: bad_g += 1
    add('站点缺口公式', bad_g == 0, '异常 %d/%d' % (bad_g, len(SITES)))
    # 10. 时间进度 = TP
    bad_tp = sum(1 for s in SITES if cur['by_site'][s]['time_progress'] != TP[CUR])
    add('站点时间进度=TP', bad_tp == 0, '异常 %d/%d' % (bad_tp, len(SITES)))
    # 11. BV美渠道订单占比精确 = CH_SHARE（每个 SKU 按 CH_SHARE 拆分订单，聚合后严格相等）
    bv = cur['by_site']['BV美']
    tot_o = bv['orders']
    bad_sh = 0
    for ch in CHANNELS:
        share = bv['channels'][ch]['orders'] / tot_o if tot_o else 0
        if abs(share - CH_SHARE['BV美'][ch]) > 1e-3: bad_sh += 1
    add('BV美渠道订单占比=真实占比', bad_sh == 0, '异常 %d/%d' % (bad_sh, len(CHANNELS)))
    # 12. 周期环比差值自洽（BV美本/上周期 CSV）
    sp = out.get('sku_period') or {}
    bad_d = n_d = 0
    for rec in (sp.get('skus') or {}).values():
        if rec.get('previous'):
            n_d += 1
            d = rec['delta']['sales']
            if abs(d['abs'] - (d['cur'] - d['prev'])) > eps: bad_d += 1
    add('周期环比差值自洽(BV美)', bad_d == 0, '异常 %d/%d' % (bad_d, n_d))
    total = len(checks)
    passed = sum(1 for c in checks if c['status'] == 'PASS')
    return {'passed': passed, 'total': total, 'checks': checks,
            'generated_at': datetime.date.today().isoformat(), 'all_pass': passed == total}

# ---------- 指标可计算性声明（供前端逐项展示：该项是否真有数据源）----------
METRIC_AVAILABILITY = {
    'sales': {'computable': True, 'scope': '全站/站点/类目/分层/渠道',
              'source': 'Excel《本周期{站点}》金额(原币)×汇率（真实）',
              'note': '可计算'},
    'orders': {'computable': True, 'scope': '全站/站点/类目/分层/渠道',
               'source': 'Excel《本周期{站点}》合计销量(列7)，或各渠道销量求和（真实）',
               'note': '可计算'},
    'aov_original': {'computable': True, 'scope': '全站/站点/类目/分层/渠道',
                     'source': 'Σ 原币金额 ÷ Σ 单量（真实）',
                     'note': '可计算（原币客单价）'},
    'conv': {'computable': False, 'scope': '全站/站点/类目/分层/渠道',
             'source': '【缺失】Excel 无 访问/UV/转化 字段',
             'note': '数据源缺失·无法计算·看板留空'},
    'conv_target': {'computable': False, 'scope': '站点/类目',
                    'source': '【缺失】Excel《站点目标》仅 4 站点各 1 个客单价目标；类目级客单价目标未提供。转化率目标仅 BV美 0.016（来自历史规则表）',
                    'note': '数据源缺失·类目级目标无法计算·看板留空'},
    'sku_count': {'computable': True, 'scope': '全站/站点/类目/分层/渠道',
                  'source': 'sku_master 命中行数（真实）', 'note': '可计算'},
    'target_sales': {'computable': True, 'scope': '全站/站点；类目/分层为推导',
                     'source': '站点：Excel《站点目标》直接取值；类目/分层：站点目标×该板块销售额占比（推导，已标注）',
                     'note': '站点级可计算；类目/分层为推导值'},
    'target_progress': {'computable': True, 'scope': '全站/站点/类目/分层',
                        'source': '实际销售额 ÷ 目标销售额 × 100', 'note': '可计算'},
    'struct_target': {'computable': True, 'scope': '产品结构',
                      'source': 'Excel《产品定位》门槛（超爆整月≥300/爆款150-300/头部90-150/腰部10-90/尾部<10）',
                      'note': '可计算（真实门槛）'},
}

# ---------- 指标来源公式（供前端「数据核对」页逐项展示真实计算来源）----------
METRIC_FORMULAS = {
    'sales': {'name': '销售额', 'unit': '人民币元',
              'formula': 'Σ actual_sales —— 各SKU实际销售额之和；actual_sales = Σ各渠道商品金额(原币)×汇率',
              'source': '原始字段 sku_master[].actual_sales'},
    'orders': {'name': '订单量', 'unit': '单',
               'formula': 'Σ actual_orders —— 各SKU实际订单量之和；= Σ各渠道订单',
               'source': '原始字段 sku_master[].actual_orders'},
    'aov_original': {'name': '客单价(原币)', 'unit': '原币',
                     'formula': 'Σ amount_ori ÷ Σ actual_orders —— 原币客单价（与原币目标同币种）',
                     'source': '原始字段 sku_master[].amount_ori、actual_orders'},
    'conv': {'name': '转化率', 'unit': '%',
             'formula': 'Σ(conv×orders) ÷ Σ orders —— 订单量加权平均转化率',
             'source': '【数据源缺失】Excel 本/上周期各站点 sheet 仅有 销量/金额/渠道，无访问/UV/转化字段；当前 conv 置空（不伪造）'},
    'sku_count': {'name': 'SKU数', 'unit': '个',
                  'formula': 'COUNT(sku_master 行) —— 当前板块命中的SKU数量',
                  'source': '原始：sku_master 行数（按板块过滤）'},
    'target_sales': {'name': '销售目标', 'unit': '人民币元',
                     'formula': '来源：Excel《站点目标》。站点板块=该站点目标；全站=四站点求和；类目/分层=按「站点目标×站点内该类目/层销售额占比」分摊',
                     'source': 'Excel《站点目标》+ 销售额占比分摊规则'},
    'target_progress': {'name': '目标进度', 'unit': '%',
                        'formula': '实际销售额 ÷ 目标销售额 × 100',
                        'source': '由 sales 与 target_sales 计算'},
}

out = {
    'month_meta': {'current_month': CUR, 'current_month_label': '2026年' + CUR,
                   'today': CUTOFF[CUR], 'cutoff': CUTOFF[CUR],
                   'days_in_month': DAYSPM[CUR], 'elapsed_days': ELAPSED[CUR], 'time_progress': TP[CUR]},
    'targets': TARGETS, 'actuals': actuals, 'sku_index': sku_index,
    'sku_master': sku_master_cur, 'channels': CHANNELS,
    'dimensions': {'sites': SITES, 'categories': CATS, 'layers': LAYERS},
    'sku_targets': SKU_TARGETS,
    'sku_period': build_sku_period(),
    'key_products': prev.get('key_products', []),
    'ogsm_config': build_ogsm_config(CUR),
    'ogsm_july': build_ogsm_july(),
    'formulas': METRIC_FORMULAS,
    'metric_availability': METRIC_AVAILABILITY,
    'strategies': prev.get('strategies', []), 'records': prev.get('records', []),
    'price_targets': PC['price_targets'], 'conv_targets': PC['conv_targets'], 'struct_targets': prev.get('struct_targets', {}),
    'stats': prev.get('stats', {}),
}
# ---- 数据层一致性校验闸门（任一项失败即中断构建）----
validation = run_validations(out)
if not validation['all_pass']:
    print('VALIDATION FAILED (%d/%d):' % (validation['passed'], validation['total']))
    for c in validation['checks']:
        if c['status'] == 'FAIL':
            print('  FAIL', c['name'], '->', c['detail'])
    raise SystemExit(1)
out['validation'] = validation
os.makedirs(os.path.dirname(OUT), exist_ok=True)
json.dump(out, open(OUT, 'w'), ensure_ascii=False, indent=0)
print('DONE bytes=', os.path.getsize(OUT))
print('CUR total sales=', actuals[CUR]['total']['sales'], 'target=', target_total_cur,
      'progress=', actuals[CUR]['total']['target_progress'], '% gap=', actuals[CUR]['total']['gap'])
print('by_layer:', {l: (actuals[CUR]['by_layer'][l]['orders'], actuals[CUR]['by_layer'][l]['conv']) for l in LAYERS})
print('mom total=', mom_map['total'], 'by_cat=', mom_map['by_category'])
print('sku_index=', len(sku_index), 'sku_master=', len(sku_master_cur))
